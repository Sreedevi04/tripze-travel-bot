import os
from datetime import datetime
from dotenv import load_dotenv
from groq import Groq
from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    filters, ContextTypes
)

load_dotenv()

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
EXCEL_FILE = os.getenv("EXCEL_FILE", "tripzy_leads.xlsx")
groq_client = Groq(api_key=os.getenv("GROQ_API_KEY"))

sessions = {}


def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(["Name", "Email", "Phone", "Telegram ID", "Destination",
                   "Days", "Travelers", "Budget", "Style", "Saved At"])
        wb.save(EXCEL_FILE)
        print(f"📄 Created: {EXCEL_FILE}")


def save_lead(session, telegram_id):
    try:
        init_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Check for existing and update
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[3].value) == str(telegram_id):
                ws.cell(i, 1, session.get("name", "Not provided"))
                ws.cell(i, 2, session.get("email", "Not provided"))
                ws.cell(i, 3, session.get("phone_contact", "Not provided"))
                ws.cell(i, 10, datetime.now().strftime("%d/%m/%Y %I:%M %p"))
                wb.save(EXCEL_FILE)
                print(f"🔄 Lead updated: {session.get('name')}")
                return
        saved_at = datetime.now().strftime("%d/%m/%Y %I:%M %p")
        ws.append([
            session.get("name", "Not provided"),
            session.get("email", "Not provided"),
            session.get("phone_contact", "Not provided"),
            str(telegram_id),
            session.get("destination", ""),
            session.get("days", ""),
            session.get("travelers", ""),
            session.get("budget", ""),
            session.get("style", ""),
            saved_at
        ])
        wb.save(EXCEL_FILE)
        print(f"💾 Lead saved: {session.get('name')} | {telegram_id}")
    except Exception as e:
        print(f"❌ Excel error: {e}")


def build_itinerary(session):
    country     = session.get("country", "India")
    destination = session.get("destination", "")
    days        = session.get("days", "")
    travelers   = session.get("travelers", "1")
    budget      = session.get("budget", "")
    style       = session.get("style", "Mix of Everything")

    prompt = f"""You are Aria, an expert travel planner. Create a detailed day-by-day itinerary.

Traveler info:
- From: {country}
- Destination: {destination}
- Duration: {days} days (Day 1 = arrival, Day {days} = departure)
- Travelers: {travelers} people
- Total Budget: {budget}
- Travel Style: {style}
- Use the local currency of {country} for all cost breakdowns.

Instructions:
- Per person budget = total budget divided by {travelers}
- Day 1 = arrival (airport transfer, check-in, light evening explore)
- Day {days} = departure (breakfast, check-out, airport drop)
- Each middle day: Morning / Afternoon / Evening sections
- Each activity: real place name + 1 line description + cost per person
- End each day with daily total per person
- Finish with Budget Summary:
  Total cost per person: X
  Total for {travelers} travelers: X
  Remaining buffer: X

Format EXACTLY like this for each day:

Day X - [Theme] [emoji]
Morning:
- [Activity] - [description] - [cost]/person
Afternoon:
- [Lunch spot] - [description] - [cost]/person
- [Activity] - [description] - [cost]/person
Evening:
- [Activity] - [description] - [cost]/person
- [Dinner spot] - [description] - [cost]/person
Daily total: [cost]/person

Use real place names. Be specific with costs. Use emojis."""

    response = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        max_tokens=3500,
        messages=[{"role": "user", "content": prompt}]
    )
    return response.choices[0].message.content

async def send_long(update: Update, text: str, chunk_size=4000):
    paragraphs = text.split("\n\n")
    current = ""
    for para in paragraphs:
        if len(current) + len(para) + 2 > chunk_size:
            if current.strip():
                await update.message.reply_text(current.strip())
            current = para
        else:
            current = (current + "\n\n" + para).strip() if current else para
    if current.strip():
        await update.message.reply_text(current.strip())



async def reply(update: Update, text: str):
    await update.message.reply_text(text)


async def process_stage(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    msg = update.message.text.strip()

    if user_id not in sessions:
        sessions[user_id] = {"stage": "greeting"}

    session = sessions[user_id]
    stage = session.get("stage", "greeting")

    print(f"\n📱 [{user_id}] Stage={stage} | Msg={msg}")

    if msg.lower() in ["hi", "hello", "start", "/start", "restart", "new trip", "hey"]:
        sessions[user_id] = {"stage": "greeting"}
        session = sessions[user_id]
        stage = "greeting"

   
    if stage == "greeting":
        session["stage"] = "ask_country"
        await reply(update,
            "✈️ Welcome to Tripzy Travel! 🌍\n\n"
            "I'm Aria, your personal AI travel planner!\n"
            "I'll build a custom day-by-day itinerary for you — just answer 6 quick questions. 🗺️"
        )
        await reply(update,
            "🌐 Question 1 of 6\n\n"
            "🏠 Which country are you currently living in?\n\n"
            "(This helps me show costs in your local currency!)"
        )

    
    elif stage == "ask_country":
        session["country"] = msg
        session["stage"] = "ask_destination"
        await reply(update,
            f"Got it! 🙌 You're from {msg}.\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"🗺️ Question 2 of 6\n\n"
            f"🌴 Where would you love to travel?\n\n"
            f"(Enter a country, city or region — e.g. Maldives, Bali, Paris)"
        )

    elif stage == "ask_destination":
        session["destination"] = msg
        session["stage"] = "ask_days"
        await reply(update,
            f"Ooh, {msg}! Amazing choice 😍\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📅 Question 3 of 6\n\n"
            f"How many days are you planning?\n\n"
            f"(Include arrival + departure days. E.g. type 7 for a 7-day trip)"
        )

    elif stage == "ask_days":
        session["days"] = msg
        session["stage"] = "ask_travelers"
        await reply(update,
            f"{msg} days — perfect! 🗓️\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"👥 Question 4 of 6\n\n"
            f"How many people are travelling?\n\n"
            f"(Just type a number. E.g. 2)"
        )

    elif stage == "ask_travelers":
        session["travelers"] = msg
        session["stage"] = "ask_budget"
        await reply(update,
            f"A trip for {msg} — fun! 🎉\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"💰 Question 5 of 6\n\n"
            f"What is your total budget for the whole trip?\n\n"
            f"(Include your currency. E.g. Rs.5,00,000 or EUR 3,000 or $5,000)"
        )

    
    elif stage == "ask_budget":
        session["budget"] = msg
        session["stage"] = "ask_style"
        await reply(update,
            f"Budget noted: {msg} 💵\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"🎨 Question 6 of 6 — Last one!\n\n"
            f"What's your travel style? Reply with a number:\n\n"
            f"1 - Adventure & Thrills\n"
            f"2 - Relaxation & Wellness\n"
            f"3 - Culture & History\n"
            f"4 - Foodie & Local Cuisine\n"
            f"5 - Mix of Everything"
        )

    elif stage == "ask_style":
        styles = {
            "1": "Adventure & Thrills",
            "2": "Relaxation & Wellness",
            "3": "Culture & History",
            "4": "Foodie & Local Cuisine",
            "5": "Mix of Everything"
        }
        session["style"] = styles.get(msg, msg)
        session["stage"] = "generating"

        await reply(update,
            f"✅ All set! Here's your trip summary:\n\n"
            f"🏠 From: {session['country']}\n"
            f"🌍 Destination: {session['destination']}\n"
            f"📅 Duration: {session['days']} days\n"
            f"👥 Travelers: {session['travelers']}\n"
            f"💰 Budget: {session['budget']}\n"
            f"🎨 Style: {session['style']}\n\n"
            f"⏳ Building your personalized itinerary... give me 15 seconds! 🪄✨"
        )

        try:
            itinerary = build_itinerary(session)
            await reply(update,
                f"🗺️ Your {session['days']}-Day {session['destination']} Itinerary ✨\n"
                f"(Sending day by day below...)"
            )
            await send_long(update, itinerary)
            session["stage"] = "confirm"
            await reply(update,
                "━━━━━━━━━━━━━━━\n"
                "💬 Happy with this itinerary?\n\n"
                "Reply YES to proceed with booking\n"
                "Reply NO to adjust the plan 😊"
            )
        except Exception as e:
            print(f"❌ Itinerary error: {e}")
            session["stage"] = "ask_style"
            await reply(update,
                "😔 Something went wrong. Please select your travel style again (1-5):"
            )

    elif stage == "confirm":
        if any(w in msg.lower() for w in ["yes", "perfect", "great", "good", "happy", "ok", "love", "sure", "proceed", "fine"]):
            session["stage"] = "ask_name"
            await reply(update, "🎉 Wonderful! Let's lock in your booking.\n\n📝 What is your full name?")
        else:
            session["stage"] = "ask_destination"
            await reply(update,
                "No worries! Let's redo the plan. 😊\n\n"
                "🗺️ What destination would you like? (You can change or keep the same)"
            )

    elif stage == "ask_name":
        session["name"] = msg
        session["stage"] = "ask_email"
        await reply(update,
            f"Lovely name, {msg}! 😊\n\n"
            f"📧 What's your email address?\n"
            f"(We'll send your full itinerary here)"
        )

    elif stage == "ask_email":
        session["email"] = msg
        session["stage"] = "ask_phone"
        await reply(update,
            "Got it! ✅\n\n"
            "📞 What's your phone number?\n"
            "(Our travel expert will call you within 24 hours)"
        )

    elif stage == "ask_phone":
        session["phone_contact"] = msg
        session["stage"] = "done"
        save_lead(session, user_id)
        await reply(update,
            f"🎊 You're all booked in, {session.get('name', '')}!\n\n"
            f"✅ Details saved\n"
            f"📞 Our expert will call you within 24 hours\n"
            f"📧 Itinerary confirmation going to: {session.get('email', '')}\n\n"
            f"Thank you for choosing Tripzy Travel! ✈️🌍\n\n"
            f"Questions? Just reply here anytime!\n"
            f"Want a new trip? Say Hi to start over."
        )

 
    elif stage == "done":
        await reply(update,
            "✈️ Your booking is already in progress!\n"
            "Our team will contact you within 24 hours. 😊\n\n"
            "Want to plan a new trip? Just say Hi!"
        )

    else:
        await reply(update,
            "I didn't quite get that 😅\n"
            "Please reply to continue, or say Hi to start over!"
        )



async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sessions[user_id] = {"stage": "greeting"}
    await process_stage(update, context)



if __name__ == "__main__":
    init_excel()
    print("✅ Tripzy Travel Bot is running on Telegram!")
    print(f"📄 Leads saving to: {EXCEL_FILE}")
    print("Press Ctrl+C to stop.\n")

    app = ApplicationBuilder().token(TELEGRAM_BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, process_stage))
    app.run_polling(close_loop=False)
